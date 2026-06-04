"""Ingestion layer for RSC season workbooks.

Parses the per-season `*_standings.xlsx` files into normalized pandas
DataFrames. Three sources matter:

  - Variables       -> per-tier config (match days, playoff cutoff, ...)
  - SnR Import       -> the full schedule + results (played and future)
  - All Teams Data   -> the league's own computed standings/metrics, which we
                        use as ground truth to validate our parser & engine.

Records are GAME-based: each match day is a 4-game series and every game
counts toward standings. A `Result` of "3-1" means 3 game-wins for Away,
1 for Home. Ranking is by game win percentage.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pandas as pd

# Canonical tier order (highest MMR -> lowest), per the RSC rulebook.
TIER_ORDER = [
    "Premier", "Master", "Elite", "Veteran", "Rival",
    "Challenger", "Prospect", "Contender", "Amateur",
]

_RESULT_RE = re.compile(r"^\s*(\d+)\s*-\s*(\d+)\s*$")


@dataclass
class Season:
    """Everything parsed out of one season workbook."""
    label: str                 # e.g. "S26"
    variables: pd.DataFrame    # one row per tier
    matches: pd.DataFrame      # one row per scheduled series
    teams: pd.DataFrame        # league's own per-team standings ("All Teams Data")

    @property
    def tiers(self) -> list[str]:
        present = set(self.matches["tier"].unique())
        return [t for t in TIER_ORDER if t in present]


def _load_wb(path: str | Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def parse_variables(wb: openpyxl.Workbook) -> pd.DataFrame:
    ws = wb["Variables"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    recs = []
    for r in rows:
        tier = r[0]
        if tier is None or tier == "Tier Name":
            continue
        recs.append({
            "tier": tier,
            "teams": _int(r[1]),
            "match_days": _int(r[2]),
            "mds_played": _int(r[3]),
            "games_left": _int(r[4]),
            "playoff_spots": _int(r[5]),
            "first_team_out": _int(r[6]),
        })
    return pd.DataFrame(recs)


def parse_matches(wb: openpyxl.Workbook) -> pd.DataFrame:
    """SnR Import -> tidy schedule.

    Columns: tier, match_day, date, away, home, away_g, home_g, total_g,
             played (bool). Future/unplayed series have played=False and
             NaN game counts.
    """
    ws = wb["SnR Import"]
    recs = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if len(r) < 8:
            r = tuple(r) + (None,) * (8 - len(r))
        match_day, date, _weekday, away, result, home, _notes, tier = r[:8]
        # Skip blank rows and repeated header rows that appear between blocks.
        if tier is None or tier == "Tier":
            continue
        if away is None or home is None or away == "Away":
            continue
        away_g = home_g = None
        played = False
        if result not in (None, "", " "):
            m = _RESULT_RE.match(str(result))
            if m:
                away_g, home_g = int(m.group(1)), int(m.group(2))
                played = True
            # non-matching results (e.g. stray "Result") are treated as unplayed
        # Regular-season match days are integers (1..N); playoff rounds are
        # labeled strings ("Wild Card", "Quarterfinals", ...). Standings count
        # the regular season only.
        md_int = _int(match_day)
        is_regular = md_int is not None
        recs.append({
            "tier": tier,
            "phase": "regular" if is_regular else str(match_day).strip(),
            "is_regular": is_regular,
            "match_day": md_int,            # int for regular season, else None
            "round_label": None if is_regular else str(match_day).strip(),
            "date": date,
            "away": str(away).strip(),
            "home": str(home).strip(),
            "away_g": away_g,
            "home_g": home_g,
            "total_g": (away_g + home_g) if played else None,
            "played": played,
        })
    df = pd.DataFrame(recs)
    return df


# Positional column map for "All Teams Data" (see ingest notes / sheet layout).
_ATD_COLS = {
    0: "tier", 1: "franchise", 2: "team", 3: "conf", 4: "division",
    5: "ov_rank", 6: "ov_wp", 7: "ov_w", 8: "ov_l", 9: "ov_gb",
    10: "cf_rank", 11: "cf_wp", 12: "cf_w", 13: "cf_l", 14: "cf_gb",
    15: "dv_rank", 16: "dv_wp", 17: "dv_w", 18: "dv_l", 19: "dv_gb",
    20: "past_owp", 21: "past_oowp", 22: "past_sos",
    23: "fut_owp", 24: "fut_oowp", 25: "fut_sos",
    26: "rpi", 27: "last5", 28: "tb", 29: "magic_number",
    30: "roster1", 31: "roster2", 32: "roster3", 33: "roster4",
}


def parse_teams(wb: openpyxl.Workbook) -> pd.DataFrame:
    """All Teams Data -> the league's own standings + computed metrics.

    This is the GROUND TRUTH we validate against. It already includes the
    league's magic number (`magic_number`) and strength-of-schedule columns.
    """
    ws = wb["All Teams Data"]
    recs = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        tier = r[0]
        if tier is None or tier == "Tier":  # skip blanks + repeated headers
            continue
        rec = {}
        for idx, name in _ATD_COLS.items():
            rec[name] = r[idx] if idx < len(r) else None
        if rec.get("team") is None:
            continue
        rec["team"] = str(rec["team"]).strip()
        recs.append(rec)
    return pd.DataFrame(recs)


def load_season(path: str | Path, label: str) -> Season:
    wb = _load_wb(path)
    return Season(
        label=label,
        variables=parse_variables(wb),
        matches=parse_matches(wb),
        teams=parse_teams(wb),
    )


def _int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


if __name__ == "__main__":
    import sys
    here = Path(__file__).resolve().parent.parent / "data"
    label = sys.argv[1] if len(sys.argv) > 1 else "S26"
    season = load_season(here / f"{label}_standings.xlsx", label)
    print(f"== {label} ==")
    print(f"tiers: {season.tiers}")
    print(f"matches: {len(season.matches)} "
          f"({int(season.matches['played'].sum())} played, "
          f"{int((~season.matches['played']).sum())} scheduled)")
    print(f"teams:  {len(season.teams)}")
    print("\nVariables:")
    print(season.variables.to_string(index=False))
